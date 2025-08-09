import torch
import os
import numpy as np

import openpyxl
from sklearn.metrics import recall_score, f1_score, accuracy_score

def confusion_matrix(matrix, conf_matrix):
    for i in matrix:
        pl,tl = i
        tl = int(tl)
        pl = int(pl)
        conf_matrix[tl, pl] = conf_matrix[tl, pl] + 1
    return conf_matrix

uar_alls = []

war_alls = []

acc_alls=[]
wf1_alls = []

uf1_alls = []

classes = ['happiness','surprise','negative']
num_class = 3
resultPath = r'.\results'
rAS_train_acc = resultPath + "\\" + 'CausalNet_CASMEII.xlsx'


matrix = torch.zeros(num_class, num_class, dtype=torch.int64)
matrix_stable = torch.zeros(num_class, num_class, dtype=torch.int64)

all_matrix=[]
all_matrix_stable=[]

all = ['sub17', 'sub26', 'sub16', 'sub09', 'sub05', 'sub24', 'sub02', 'sub13', 'sub04', 'sub23', 'sub11', 'sub12', 'sub08', 'sub14', 'sub03', 'sub19', 'sub01',
            'sub20', 'sub21', 'sub22', 'sub15', 'sub06', 'sub25', 'sub07']
for i in all:
    with open(os.path.join('.', 'results', str(i)+'_acc.txt'), "r") as f:
        data = f.readlines()
        matrix_tmp = []
        tmp = data[2][12:-2]
        tmp2 = tmp.split(', ')
        tmpnum1 = ''
        tmpnum2 = ''
        for j in range(0, len(tmp2)):
            if j % 2 == 0:
                for l in tmp2[j]:
                    if l >= '0' and l <= '9':
                        tmpnum1 = tmpnum1 + l
            else:
                for l in tmp2[j]:
                    if l >= '0' and l <= '9':
                        tmpnum2 = tmpnum2 + l
                matrix_tmp.append([int(tmpnum1), int(tmpnum2)])
                tmpnum1 = ''
                tmpnum2 = ''
        all_matrix = all_matrix + matrix_tmp


# print(all_matrix)
y_true=[]
y_pred=[]
for l in range(0,len(all_matrix)):
    y_pred.append(all_matrix[l][0])
    y_true.append(all_matrix[l][1])
uar = recall_score(y_true, y_pred, average='macro')
war = recall_score(y_true, y_pred, average='weighted')
f1 = f1_score(y_true, y_pred, average='weighted')
uf1 = f1_score(y_true, y_pred, average='macro')
acc = accuracy_score(y_true, y_pred)

matrix = confusion_matrix(all_matrix, matrix)



acc_alls.append(acc)

uar_alls.append(uar)

war_alls.append(war)

uf1_alls.append(uf1)

wf1_alls.append(f1)

# 结果写入到excel中
excel_o = openpyxl.Workbook()
sheet_o = excel_o.active
# 保存最好结果到excel
sheet_o.cell(row=1, column=1, value='col is pre')
for i_row in range(num_class):
    sheet_o.cell(row=1, column=i_row + 2, value=classes[i_row])
    sheet_o.cell(row=i_row + 2, column=1, value=classes[i_row])
    for i_col in range(num_class):
        sheet_o.cell(row=i_row + 2, column=i_col + 2, value=matrix[i_row][i_col].item())
sheet_o.cell(row=1, column=num_class + 3, value='all_war')
sheet_o.cell(row=1, column=num_class + 4, value='all_uar')
sheet_o.cell(row=1, column=num_class + 5, value='all_f1')
sheet_o.cell(row=1, column=num_class + 6, value='all_uf1')
sheet_o.cell(row=1, column=num_class + 7, value='all_acc')
sheet_o.cell(row=2, column=num_class + 3, value=war_alls[0])
sheet_o.cell(row=2, column=num_class + 4, value=uar_alls[0])
sheet_o.cell(row=2, column=num_class + 5, value=wf1_alls[0])
sheet_o.cell(row=2, column=num_class + 6, value=uf1_alls[0])
sheet_o.cell(row=2, column=num_class + 7, value=acc_alls[0])


excel_o.save(rAS_train_acc)
