import torch
import os


import openpyxl
from sklearn.metrics import recall_score, f1_score, accuracy_score

def confusion_matrix(matrix, conf_matrix):
    for i in matrix:
        pl,tl = i
        tl = int(tl)
        pl = int(pl)
        conf_matrix[tl, pl] = conf_matrix[tl, pl] + 1
    return conf_matrix

uar_alls = []

war_alls = []

acc_alls=[]
wf1_alls = []

uf1_alls = []

classes = ['happiness','surprise','negative']
num_class = 3
resultPath = r'.\results'
rAS_train_acc = resultPath + "\\" + 'CausalNet_SAMM.xlsx'


matrix = torch.zeros(num_class, num_class, dtype=torch.int64)
matrix_stable = torch.zeros(num_class, num_class, dtype=torch.int64)
# 计算混淆矩阵
all_matrix=[]
all_matrix_stable=[]

all = ['006','007','009','010','011','012','013','014','015','016','017','018','019','020','021','022','023','024','026','028','030','032','031','033','034','035','036','037']
for i in all:
    with open(os.path.join('.', 'results', str(i)+'_acc.txt'), "r") as f:
        data = f.readlines()
        matrix_tmp = []
        tmp = data[2][12:-2]
        tmp2 = tmp.split(', ')
        tmpnum1 = ''
        tmpnum2 = ''
        for j in range(0, len(tmp2)):
            if j % 2 == 0:
                for l in tmp2[j]:
                    if l >= '0' and l <= '9':
                        tmpnum1 = tmpnum1 + l
            else:
                for l in tmp2[j]:
                    if l >= '0' and l <= '9':
                        tmpnum2 = tmpnum2 + l
                matrix_tmp.append([int(tmpnum1), int(tmpnum2)])
                tmpnum1 = ''
                tmpnum2 = ''
        all_matrix = all_matrix + matrix_tmp


# print(all_matrix)
y_true=[]
y_pred=[]
for l in range(0,len(all_matrix)):
    y_pred.append(all_matrix[l][0])
    y_true.append(all_matrix[l][1])
uar = recall_score(y_true, y_pred, average='macro')
war = recall_score(y_true, y_pred, average='weighted')
f1 = f1_score(y_true, y_pred, average='weighted')
uf1 = f1_score(y_true, y_pred, average='macro')
acc = accuracy_score(y_true, y_pred)

matrix = confusion_matrix(all_matrix, matrix)



acc_alls.append(acc)

uar_alls.append(uar)

war_alls.append(war)

uf1_alls.append(uf1)

wf1_alls.append(f1)


excel_o = openpyxl.Workbook()
sheet_o = excel_o.active

sheet_o.cell(row=1, column=1, value='col is pre')
for i_row in range(num_class):
    sheet_o.cell(row=1, column=i_row + 2, value=classes[i_row])
    sheet_o.cell(row=i_row + 2, column=1, value=classes[i_row])
    for i_col in range(num_class):
        sheet_o.cell(row=i_row + 2, column=i_col + 2, value=matrix[i_row][i_col].item())
sheet_o.cell(row=1, column=num_class + 3, value='all_war')
sheet_o.cell(row=1, column=num_class + 4, value='all_uar')
sheet_o.cell(row=1, column=num_class + 5, value='all_f1')
sheet_o.cell(row=1, column=num_class + 6, value='all_uf1')
sheet_o.cell(row=1, column=num_class + 7, value='all_acc')
sheet_o.cell(row=2, column=num_class + 3, value=war_alls[0])
sheet_o.cell(row=2, column=num_class + 4, value=uar_alls[0])
sheet_o.cell(row=2, column=num_class + 5, value=wf1_alls[0])
sheet_o.cell(row=2, column=num_class + 6, value=uf1_alls[0])
sheet_o.cell(row=2, column=num_class + 7, value=acc_alls[0])


excel_o.save(rAS_train_acc)
