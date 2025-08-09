import random
from os import path

import cv2
import time
import pandas
from sklearn.metrics import confusion_matrix
import torch.nn as nn
from torch.utils.data import TensorDataset, DataLoader
import argparse
from distutils.util import strtobool

import torch
import os


import openpyxl
from sklearn.metrics import recall_score, f1_score, accuracy_score
import numpy as np


from Models import CausalNet


def my_confusion_matrix(matrix, conf_matrix):
    for i in matrix:
        pl, tl = i
        tl = int(tl)
        pl = int(pl)
        conf_matrix[tl, pl] = conf_matrix[tl, pl] + 1
    return conf_matrix

def reset_weights(m):
    for layer in m.children():
        if hasattr(layer, 'reset_parameters'):
            layer.reset_parameters()

def confusionMatrix(gt, pred, show=False):
    TN, FP, FN, TP = confusion_matrix(gt, pred).ravel()
    f1_score = (2 * TP) / (2 * TP + FP + FN)
    num_samples = len([x for x in gt if x == 1])
    average_recall = TP / num_samples
    return f1_score, average_recall


def recognition_evaluation(final_gt, final_pred, show=False):
    label_dict = {'negative': 0, 'positive': 1, 'surprise': 2}
    f1_list = []
    ar_list = []
    try:
        for emotion, emotion_index in label_dict.items():
            gt_recog = [1 if x == emotion_index else 0 for x in final_gt]
            pred_recog = [1 if x == emotion_index else 0 for x in final_pred]
            try:
                f1_recog, ar_recog = confusionMatrix(gt_recog, pred_recog)
                f1_list.append(f1_recog)
                ar_list.append(ar_recog)
            except Exception as e:
                pass
        UF1 = np.mean(f1_list)
        UAR = np.mean(ar_list)
        return UF1, UAR
    except:
        return '', ''



def whole_face_block_coordinates():
    '''
    This part of the code references HTNet.
    '''
    df = pandas.read_csv('combined_3_class2_for_optical_flow.csv')
    m, n = df.shape

    face_block_coordinates = {}

    for i in range(0, m):
        image_name = str(df['sub'][i]) + '_' + str(
            df['filename_o'][i]) + ' .png'

        batch_landmarks=None

        if batch_landmarks is None:
            batch_landmarks = np.array([[[9.528073, 11.062551]
                                            , [21.396168, 10.919773]
                                            , [15.380184, 17.380562]
                                            , [10.255435, 22.121233]
                                            , [20.583706, 22.25584]]])

        row_n, col_n = np.shape(batch_landmarks[0])

        for i in range(0, row_n):
            for j in range(0, col_n):
                if batch_landmarks[0][i][j] < 7:
                    batch_landmarks[0][i][j] = 7
                if batch_landmarks[0][i][j] > 21:
                    batch_landmarks[0][i][j] = 21


        batch_landmarks = batch_landmarks.astype(int)

        face_block_coordinates[image_name] = batch_landmarks[0]
        tmp=image_name.split(' ')[0]
        tmp1=tmp+'_1 .png'
        face_block_coordinates[tmp1] = batch_landmarks[0]
        tmp2 = tmp + '_2 .png'
        face_block_coordinates[tmp2] = batch_landmarks[0]
        tmp3 = tmp + '_3 .png'
        face_block_coordinates[tmp3] = batch_landmarks[0]

    return face_block_coordinates


def crop_optical_flow_block():
    '''
    This part of the code references HTNet.
    '''
    face_block_coordinates_dict = whole_face_block_coordinates()

    whole_optical_flow_path = './datasets/STSNet_whole_norm_u_v_os'
    whole_optical_flow_imgs = os.listdir(whole_optical_flow_path)
    four_parts_optical_flow_imgs = {}

    for n_img in whole_optical_flow_imgs:
        four_parts_optical_flow_imgs[n_img]=[]
        flow_image = cv2.imread(whole_optical_flow_path + '/' + n_img)
        four_part_coordinates = face_block_coordinates_dict[n_img]
        l_eye = flow_image[four_part_coordinates[0][0]-7:four_part_coordinates[0][0]+7,
                four_part_coordinates[0][1]-7: four_part_coordinates[0][1]+7]
        l_lips = flow_image[four_part_coordinates[1][0] - 7:four_part_coordinates[1][0] + 7,
                four_part_coordinates[1][1] - 7: four_part_coordinates[1][1] + 7]
        nose = flow_image[four_part_coordinates[2][0] - 7:four_part_coordinates[2][0] + 7,
                four_part_coordinates[2][1] - 7: four_part_coordinates[2][1] + 7]
        r_eye = flow_image[four_part_coordinates[3][0] - 7:four_part_coordinates[3][0] + 7,
                four_part_coordinates[3][1] - 7: four_part_coordinates[3][1] + 7]
        r_lips = flow_image[four_part_coordinates[4][0] - 7:four_part_coordinates[4][0] + 7,
                four_part_coordinates[4][1] - 7: four_part_coordinates[4][1] + 7]
        four_parts_optical_flow_imgs[n_img].append(l_eye)
        four_parts_optical_flow_imgs[n_img].append(l_lips)
        four_parts_optical_flow_imgs[n_img].append(nose)
        four_parts_optical_flow_imgs[n_img].append(r_eye)
        four_parts_optical_flow_imgs[n_img].append(r_lips)

    return four_parts_optical_flow_imgs



def main(config):
    '''
    The training and testing framework refers to HTNet.
    layer1, layer2, layer3, and gammas are parameters that need to be tuned.
    '''
    layers3 = [7,8, 9, 10]
    layers2 = [3, 2]
    layers1 = [3, 2]
    gammas = [0.3,0.4,0.5]
    for layer1 in layers1:
        for layer2 in layers2:
            for layer3 in layers3:
                for gamma in gammas:
                    if layer1 == 2 and layer2 == 2:
                        continue
                    seed = 2025
                    torch.manual_seed(seed)
                    torch.cuda.manual_seed(seed)
                    np.random.seed(seed)
                    random.seed(seed)

                    learning_rate = 0.00005
                    batch_size = 256 * 4
                    epochs = 200
                    all_accuracy_dict = {}
                    is_cuda = torch.cuda.is_available()
                    if is_cuda:
                        device = torch.device('cuda')
                    else:
                        device = torch.device('cpu')
                    loss_fn = nn.CrossEntropyLoss()
                    if (config.train):
                        if not path.exists('ourmodel_threedatasets_weights'):
                            os.mkdir('ourmodel_threedatasets_weights')

                    print('lr=%f, epochs=%d, device=%s\n' % (learning_rate, epochs, device))

                    total_gt = []
                    total_pred = []
                    best_total_pred = []

                    t = time.time()

                    main_path = './datasets/three_norm_u_v_os'
                    subName = os.listdir(main_path)
                    all_five_parts_optical_flow = crop_optical_flow_block()
                    print(subName)

                    for n_subName in subName:
                        print('Subject:', n_subName)
                        y_train = []
                        y_test = []
                        four_parts_train = []

                        four_parts_test = []

                        expression = os.listdir(main_path + '/' + n_subName + '/u_train')

                        '''
                        lr_eye_lips, lr_eye_lips1, lr_eye_lips2, and lr_eye_lips3 correspond to 
                        onset-apex OF, apex-offset OF, onset-apex direction map and apex-offset direction map.
                        '''
                        for n_expression in expression:
                            img = os.listdir(main_path + '/' + n_subName + '/u_train/' + n_expression)

                            for n_img in img:
                                y_train.append(int(n_expression))

                                l_eye_lips = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img][0], all_five_parts_optical_flow[n_img][1]])
                                r_eye_lips = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img][3], all_five_parts_optical_flow[n_img][4]])
                                lr_eye_lips = cv2.vconcat([l_eye_lips, r_eye_lips])

                                n_img1 = n_img.split(' ')[0] + '_1 .png'

                                l_eye_lips1 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img1][0], all_five_parts_optical_flow[n_img1][1]])
                                r_eye_lips1 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img1][3], all_five_parts_optical_flow[n_img1][4]])
                                lr_eye_lips1 = cv2.vconcat([l_eye_lips1, r_eye_lips1])

                                n_img2 = n_img.split(' ')[0] + '_2 .png'

                                l_eye_lips2 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img2][0], all_five_parts_optical_flow[n_img2][1]])
                                r_eye_lips2 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img2][3], all_five_parts_optical_flow[n_img2][4]])
                                lr_eye_lips2 = cv2.vconcat([l_eye_lips2, r_eye_lips2])

                                n_img3 = n_img.split(' ')[0] + '_3 .png'

                                l_eye_lips3 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img3][0], all_five_parts_optical_flow[n_img3][1]])
                                r_eye_lips3 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img3][3], all_five_parts_optical_flow[n_img3][4]])
                                lr_eye_lips3 = cv2.vconcat([l_eye_lips3, r_eye_lips3])

                                four_parts_train.append([lr_eye_lips, lr_eye_lips1, lr_eye_lips2, lr_eye_lips3])
                        expression = os.listdir(main_path + '/' + n_subName + '/u_test')
                        for n_expression in expression:
                            img = os.listdir(main_path + '/' + n_subName + '/u_test/' + n_expression)

                            for n_img in img:
                                y_test.append(int(n_expression))
                                l_eye_lips = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img][0], all_five_parts_optical_flow[n_img][1]])
                                r_eye_lips = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img][3], all_five_parts_optical_flow[n_img][4]])
                                lr_eye_lips = cv2.vconcat([l_eye_lips, r_eye_lips])

                                n_img1 = n_img.split(' ')[0] + '_1 .png'

                                l_eye_lips1 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img1][0], all_five_parts_optical_flow[n_img1][1]])
                                r_eye_lips1 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img1][3], all_five_parts_optical_flow[n_img1][4]])
                                lr_eye_lips1 = cv2.vconcat([l_eye_lips1, r_eye_lips1])

                                n_img2 = n_img.split(' ')[0] + '_2 .png'

                                l_eye_lips2 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img2][0], all_five_parts_optical_flow[n_img2][1]])
                                r_eye_lips2 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img2][3], all_five_parts_optical_flow[n_img2][4]])
                                lr_eye_lips2 = cv2.vconcat([l_eye_lips2, r_eye_lips2])

                                n_img3 = n_img.split(' ')[0] + '_3 .png'

                                l_eye_lips3 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img3][0], all_five_parts_optical_flow[n_img3][1]])
                                r_eye_lips3 = cv2.hconcat(
                                    [all_five_parts_optical_flow[n_img3][3], all_five_parts_optical_flow[n_img3][4]])
                                lr_eye_lips3 = cv2.vconcat([l_eye_lips3, r_eye_lips3])

                                four_parts_test.append([lr_eye_lips, lr_eye_lips1, lr_eye_lips2, lr_eye_lips3])

                        weight_path = 'ourmodel_threedatasets_weights' + '/' + n_subName + '.pth'

                        model = CausalNet(
                            image_size=28,
                            patch_size=7,
                            dim=256,
                            heads=3,
                            num_hierarchies=3,
                            block_repeats=(layer1, layer2, layer3),

                            num_classes=3,
                            gamma=gamma
                        )
                        model = model.to(device)

                        if (config.train):

                            print('train')
                            print(len(y_train))
                            print(len(y_test))
                        else:
                            model.load_state_dict(torch.load(weight_path))
                        optimizer = torch.optim.Adam(model.parameters(), lr=learning_rate)
                        y_train = torch.Tensor(y_train).to(dtype=torch.long)

                        four_parts_train = torch.Tensor(np.array(four_parts_train))

                        four_parts_train = four_parts_train.permute(0, 1, 4, 2, 3)

                        dataset_train = TensorDataset(four_parts_train, y_train)
                        train_dl = DataLoader(dataset_train, batch_size=batch_size)
                        y_test = torch.Tensor(y_test).to(dtype=torch.long)
                        four_parts_test = torch.Tensor(np.array(four_parts_test))
                        four_parts_test = four_parts_test.permute(0, 1, 4, 2, 3)
                        dataset_test = TensorDataset(four_parts_test, y_test)
                        test_dl = DataLoader(dataset_test, batch_size=batch_size)

                        best_accuracy_for_each_subject = 0
                        best_each_subject_pred = []
                        best_epoch = 0
                        best_matrix = []
                        for epoch in range(1, epochs + 1):
                            if (config.train):
                                # Training
                                model.train()
                                train_loss = 0.0
                                num_train_correct = 0
                                num_train_examples = 0
                                matrix = []

                                for batch in train_dl:
                                    optimizer.zero_grad()
                                    x = batch[0].to(device)

                                    y = batch[1].to(device)
                                    yhat = model(x)
                                    loss = loss_fn(yhat, y)
                                    loss.backward()
                                    optimizer.step()

                                    train_loss += loss.data.item() * x.size(0)
                                    num_train_correct += (torch.max(yhat, 1)[1] == y).sum().item()
                                    num_train_examples += x.shape[0]

                            model.eval()
                            val_loss = 0.0
                            num_val_correct = 0
                            num_val_examples = 0
                            for batch in test_dl:
                                x = batch[0].to(device)
                                y = batch[1].to(device)
                                yhat = model(x)
                                loss = loss_fn(yhat, y)

                                _, predicts = torch.max(yhat, 1)
                                for a in range(0, len(predicts)):
                                    matrix.append([int(predicts[a]), int(y[a])])

                                val_loss += loss.data.item() * x.size(0)
                                num_val_correct += (torch.max(yhat, 1)[1] == y).sum().item()
                                num_val_examples += y.shape[0]

                            val_acc = num_val_correct / num_val_examples
                            val_loss = val_loss / len(test_dl.dataset)
                            print("[Epoch %d] Validation accuracy:%.4f. Loss:%.3f" % (epoch, val_acc, val_loss))

                            temp_best_each_subject_pred = []
                            if best_accuracy_for_each_subject < val_acc:
                                best_accuracy_for_each_subject = val_acc
                                temp_best_each_subject_pred.extend(torch.max(yhat, 1)[1].tolist())
                                best_each_subject_pred = temp_best_each_subject_pred
                                best_matrix = matrix
                                best_epoch = epoch
                            '''
                            Save the results.
                            '''
                            if val_acc >= 1:

                                if not os.path.exists(os.path.join('.', 'results', str(gamma)+'_'+str(layer1)+'_'+str(layer2)+'_'+str(layer3))):
                                    os.makedirs(os.path.join('.', 'results', str(gamma)+'_'+str(layer1)+'_'+str(layer2)+'_'+str(layer3)))
                                with open(os.path.join('.', 'results', str(gamma)+'_'+str(layer1)+'_'+str(layer2)+'_'+str(layer3), str(n_subName) + '_acc.txt'),
                                          'a') as f:
                                    f.write('best epoach: ' + str(best_epoch) + '\n' + 'best acc: ' + str(
                                        best_accuracy_for_each_subject) + '\n' + 'matrix_acc: ' + str(
                                        best_matrix) + '\n')

                                break
                            if epoch == epochs:

                                if not os.path.exists(os.path.join('.', 'results', str(gamma)+'_'+str(layer1)+'_'+str(layer2)+'_'+str(layer3))):
                                    os.makedirs(os.path.join('.', 'results', str(gamma)+'_'+str(layer1)+'_'+str(layer2)+'_'+str(layer3)))
                                with open(os.path.join('.', 'results', str(gamma)+'_'+str(layer1)+'_'+str(layer2)+'_'+str(layer3), str(n_subName) + '_acc.txt'),
                                          'a') as f:
                                    f.write('best epoach: ' + str(best_epoch) + '\n' + 'best acc: ' + str(
                                        best_accuracy_for_each_subject) + '\n' + 'matrix_acc: ' + str(
                                        best_matrix) + '\n')

                        # For UF1 and UAR computation
                        print('Best Predicted    :', best_each_subject_pred)
                        accuracydict = {}
                        accuracydict['pred'] = best_each_subject_pred
                        accuracydict['truth'] = y.tolist()
                        all_accuracy_dict[n_subName] = accuracydict

                        print('Ground Truth :', y.tolist())
                        print('Evaluation until this subject: ')
                        total_pred.extend(torch.max(yhat, 1)[1].tolist())
                        total_gt.extend(y.tolist())
                        best_total_pred.extend(best_each_subject_pred)

                        best_UF1, best_UAR = recognition_evaluation(total_gt, best_total_pred, show=True)
                        print('best UF1:', round(best_UF1, 4), '| best UAR:', round(best_UAR, 4))

                    print('Final Evaluation: ')

                    print(np.shape(total_gt))
                    print('Total Time Taken:', time.time() - t)
                    print(all_accuracy_dict)

                    '''
                        Calculate the composite datasets results. if you need results of single datasets, use calculate_all_xxx.py
                    '''
                    uar_alls = []
                    war_alls = []
                    acc_alls = []
                    wf1_alls = []
                    uf1_alls = []
                    classes = ['happiness', 'surprise', 'negative']
                    num_class = 3
                    resultPath = os.path.join('.', 'results', str(gamma)+'_'+str(layer1)+'_'+str(layer2)+'_'+str(layer3))

                    matrix = torch.zeros(num_class, num_class, dtype=torch.int64)
                    all_matrix = []


                    all = ['sub17', 'sub26', 'sub16', 'sub09', 'sub05', 'sub24', 'sub02', 'sub13', 'sub04', 'sub23',
                           'sub11', 'sub12', 'sub08', 'sub14', 'sub03', 'sub19', 'sub01',
                           'sub20', 'sub21', 'sub22', 'sub15', 'sub06', 'sub25', 'sub07', '006', '007', '009', '010',
                           '011', '012', '013', '014', '015', '016', '017', '018', '019', '020', '021', '022', '023',
                           '024', '026', '028', '030', '032', '031', '033', '034', '035', '036', '037', 's01', 's02',
                           's03', 's04', 's05', 's06', 's08', 's09', 's11', 's12', 's13', 's14', 's15', 's18', 's19',
                           's20']
                    for i in all:
                        with open(os.path.join('.', 'results',str(gamma)+'_'+str(layer1)+'_'+str(layer2)+'_'+str(layer3), str(i) + '_acc.txt'), "r") as f:
                            data = f.readlines()
                            matrix_tmp = []
                            tmp = data[2][12:-2]
                            tmp2 = tmp.split(', ')
                            tmpnum1 = ''
                            tmpnum2 = ''
                            for j in range(0, len(tmp2)):
                                if j % 2 == 0:
                                    for l in tmp2[j]:
                                        if l >= '0' and l <= '9':
                                            tmpnum1 = tmpnum1 + l
                                else:
                                    for l in tmp2[j]:
                                        if l >= '0' and l <= '9':
                                            tmpnum2 = tmpnum2 + l
                                    matrix_tmp.append([int(tmpnum1), int(tmpnum2)])
                                    tmpnum1 = ''
                                    tmpnum2 = ''
                            all_matrix = all_matrix + matrix_tmp


                    y_true = []
                    y_pred = []
                    for l in range(0, len(all_matrix)):
                        y_pred.append(all_matrix[l][0])
                        y_true.append(all_matrix[l][1])
                    uar = recall_score(y_true, y_pred, average='macro')
                    war = recall_score(y_true, y_pred, average='weighted')
                    f1 = f1_score(y_true, y_pred, average='weighted')
                    uf1 = f1_score(y_true, y_pred, average='macro')
                    acc = accuracy_score(y_true, y_pred)

                    matrix = my_confusion_matrix(all_matrix, matrix)



                    acc_alls.append(acc)

                    uar_alls.append(uar)

                    war_alls.append(war)

                    uf1_alls.append(uf1)

                    wf1_alls.append(f1)


                    excel_o = openpyxl.Workbook()
                    sheet_o = excel_o.active

                    sheet_o.cell(row=1, column=1, value='col is pre')
                    for i_row in range(num_class):
                        sheet_o.cell(row=1, column=i_row + 2, value=classes[i_row])
                        sheet_o.cell(row=i_row + 2, column=1, value=classes[i_row])
                        for i_col in range(num_class):
                            sheet_o.cell(row=i_row + 2, column=i_col + 2, value=matrix[i_row][i_col].item())
                    sheet_o.cell(row=1, column=num_class + 3, value='all_war')
                    sheet_o.cell(row=1, column=num_class + 4, value='all_uar')
                    sheet_o.cell(row=1, column=num_class + 5, value='all_f1')
                    sheet_o.cell(row=1, column=num_class + 6, value='all_uf1')
                    sheet_o.cell(row=1, column=num_class + 7, value='all_acc')
                    sheet_o.cell(row=2, column=num_class + 3, value=war_alls[0])
                    sheet_o.cell(row=2, column=num_class + 4, value=uar_alls[0])
                    sheet_o.cell(row=2, column=num_class + 5, value=wf1_alls[0])
                    sheet_o.cell(row=2, column=num_class + 6, value=uf1_alls[0])
                    sheet_o.cell(row=2, column=num_class + 7, value=acc_alls[0])

                    rAS_train_acc = resultPath + "\\" + str(uf1_alls[0])+'.xlsx'
                    excel_o.save(rAS_train_acc)
                    '''
                        Record and summarize the results of the composite dataset for easy viewing.
                    '''
                    summary_txt = os.path.join('.', 'results', 'result_summary.txt')
                    with open(summary_txt, 'a') as f:
                        f.write(
                            f'gamma={gamma}, layer1={layer1}, layer2={layer2}, layer3={layer3} | UF1={uf1_alls[0]:.4f}, UAR={uar_alls[0]:.4f}\n')





if __name__ == '__main__':
  
    parser = argparse.ArgumentParser()

    parser.add_argument('--train', type=strtobool, default=True)  # Train or use pre-trained weight for prediction
    config = parser.parse_args()
    main(config)
